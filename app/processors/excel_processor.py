# app/processors/excel_processor.py
import io
import time
import logging
from openpyxl import load_workbook
from app.processors.base_processor import BaseProcessor
from app.utils.memory_utils import cleanup_variables, clear_memory

logger = logging.getLogger(__name__)

class ExcelProcessor(BaseProcessor):
    """Excel document processor with proper header-value association"""
    
    def get_supported_extensions(self) -> set:
        return {'xlsx', 'xls'}
    
    def supports_format(self, file_format: str) -> bool:
        """Check if the processor supports the given file format"""
        return file_format.lower() in self.get_supported_extensions()
    
    async def process(self, file_content: bytes, filename: str, **kwargs) -> str:
        """Parse Excel file with proper header-value association"""
        logger.info("Starting Excel parsing with header association")
        start_time = time.time()
        
        try:
            # Load workbook from bytes
            workbook = load_workbook(io.BytesIO(file_content), data_only=True)
            all_text_parts = []
            
            for sheet_name in workbook.sheetnames:
                logger.info(f"Processing Excel sheet: {sheet_name}")
                sheet = workbook[sheet_name]
                
                # Get sheet dimensions efficiently
                max_row = sheet.max_row
                max_col = sheet.max_column
                
                if max_row == 1 or max_col == 1:
                    continue  # Skip empty or single-cell sheets
                    
                # Convert to list of lists for easier processing
                data_rows = []
                for row in sheet.iter_rows(min_row=1, max_row=min(max_row, 1000), values_only=True):  # Limit rows for memory
                    # Convert None to empty string and clean data
                    cleaned_row = [str(cell).strip() if cell is not None else "" for cell in row]
                    if any(cleaned_row):  # Only add non-empty rows
                        data_rows.append(cleaned_row)
                
                if len(data_rows) < 2:
                    continue  # Need at least header + 1 data row
                    
                # Identify header row (usually first row)
                headers = data_rows[0]
                data_rows = data_rows[1:]
                
                # Clean and validate headers
                clean_headers = []
                for i, header in enumerate(headers):
                    if header and str(header).strip():
                        clean_headers.append(str(header).strip())
                    else:
                        clean_headers.append(f"Column_{i+1}")  # Default name for empty headers
                
                # Build structured text with header-value pairs
                sheet_text_parts = [f"=== EXCEL SHEET: {sheet_name} ===\n"]
                
                # Method 1: Row-wise representation with clear header-value mapping
                for row_idx, row_data in enumerate(data_rows[:500], 1):  # Limit rows for memory
                    if not any(str(cell).strip() for cell in row_data if cell):  # Skip empty rows
                        continue
                        
                    row_text_parts = [f"Row {row_idx}:"]
                    for header, value in zip(clean_headers, row_data):
                        if value and str(value).strip():
                            clean_value = str(value).strip()
                            # Create clear associations
                            row_text_parts.append(f"{header}: {clean_value}")
                    
                    if len(row_text_parts) > 1:  # Has actual data
                        sheet_text_parts.append(" | ".join(row_text_parts))
                
                # Method 2: Column-wise summary for better context
                sheet_text_parts.append(f"\n--- COLUMN SUMMARIES for {sheet_name} ---")
                for col_idx, header in enumerate(clean_headers):
                    if col_idx >= len(data_rows[0]) if data_rows else True:
                        continue
                        
                    # Extract column values
                    column_values = []
                    for row in data_rows[:100]:  # Sample first 100 rows
                        if col_idx < len(row) and row[col_idx] and str(row[col_idx]).strip():
                            val = str(row[col_idx]).strip()
                            if val not in column_values:  # Avoid duplicates
                                column_values.append(val)
                            if len(column_values) >= 10:  # Limit unique values shown
                                break
                    
                    if column_values:
                        values_preview = ", ".join(column_values[:5])
                        if len(column_values) > 5:
                            values_preview += f"... ({len(column_values)} unique values)"
                        sheet_text_parts.append(f"{header} contains: {values_preview}")
                
                all_text_parts.extend(sheet_text_parts)
                all_text_parts.append("\n" + "="*50 + "\n")
                
                # Memory cleanup with enhanced logging
                cleanup_variables(data_rows)
                clear_memory(f"Excel sheet {sheet_name}")
            
            workbook.close()
            final_text = "\n".join(all_text_parts)
            
            logger.info(f"Excel parsing completed in {time.time() - start_time:.2f}s. Text length: {len(final_text)}")
            return final_text
            
        except Exception as e:
            logger.error(f"Excel parsing failed: {str(e)}")
            raise Exception(f"Excel parsing failed: {str(e)}")